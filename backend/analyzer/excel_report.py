from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO

def create_excel_report(data, drawing_type, window_tag_counts=None, flag_details=None, windows_only=False):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{drawing_type.capitalize()} Summary"

    # Collect all unique window tags
    unique_tags = set()
    for entry in data:
        unique_tags.update(entry.get("window_types", {}).keys())

    sorted_tags = sorted(unique_tags)

    # Build header row
    if windows_only:
        header = ["Page #"] + sorted_tags + ["Total"]
    else:
        header = ["Page #"] + sorted_tags + ["Total Windows", "Total Doors"]
    ws.append(header)

    # Format header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Populate data rows
    for entry in data:
        page_label = entry.get("printed_page") or entry.get("page") or "Unknown"
        row = [page_label]
        page_total = 0

        for tag in sorted_tags:
            count = entry.get("window_types", {}).get(tag, 0)
            row.append(count)
            page_total += count

        row.append(page_total)

        if not windows_only:
            row.append(entry.get("doors", 0))

        ws.append(row)

    # Add totals row
    total_row = ["TOTAL"]
    for tag in sorted_tags:
        total_count = sum(entry.get("window_types", {}).get(tag, 0) for entry in data)
        total_row.append(total_count)

    total_row.append(sum(sum(entry.get("window_types", {}).values()) for entry in data))
    if not windows_only:
        total_row.append(sum(entry.get("doors", 0) for entry in data))

    ws.append(total_row)

    # Center all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Optional: create flag sheet
    if flag_details:
        flags_ws = wb.create_sheet("Flags Report")
        flags_ws.append(["Page", "Tile", "Flag Type", "Details"])

        # Categorize flags
        flag_categories = {
            "window": [],
            "door": [],
            "other": []
        }

        for flag in flag_details:
            flag_text = flag.get("Flag") or flag.get("Reason") or ""
            flag_lower = flag_text.lower()
            if "window" in flag_lower:
                flag_categories["window"].append(flag)
            elif "door" in flag_lower:
                flag_categories["door"].append(flag)
            else:
                flag_categories["other"].append(flag)

        # Write categorized flags
        for category, items in flag_categories.items():
            if items:
                flags_ws.append([f"--- {category.upper()} FLAGS ---", "", "", ""])
                for flag in items:
                    flags_ws.append([
                        flag.get("Page", "Unknown"),
                        flag.get("Tile", "Unknown"),
                        category.upper(),
                        flag.get("Flag") or flag.get("Reason", "N/A")
                    ])

        # Format flag sheet
        for row in flags_ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Save to memory
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
